"""Excel validation report writer."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from segy_toolbox.models import BatchResult, ValidationResult


# Cell styling
_HEADER_FILL = PatternFill(start_color="2B3E50", end_color="2B3E50", fill_type="solid")
_HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
_PASS_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
_FAIL_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
_WARN_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
_STATUS_FILLS = {"PASS": _PASS_FILL, "FAIL": _FAIL_FILL, "WARNING": _WARN_FILL}


def write_validation_report(
    results: list[BatchResult],
    output_path: str | Path,
) -> str:
    """Write an Excel report with summary and per-file validation details."""
    output_path = str(output_path)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # --- Summary sheet ---
        summary_data = []
        for r in results:
            pre_status = r.validation_before.overall_status if r.validation_before else "N/A"
            post_status = r.validation_after.overall_status if r.validation_after else "N/A"
            summary_data.append({
                "File": r.filename,
                "Status": r.status,
                "Pre-Validation": pre_status,
                "Post-Validation": post_status,
                "Changes": len(r.changes),
                "Duration (s)": f"{r.duration_seconds:.1f}",
                "Message": r.message,
            })

        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

        # --- Per-file validation detail sheets ---
        for r in results:
            val_result = r.validation_before or r.validation_after
            if not val_result:
                continue

            sheet_name = _safe_sheet_name(r.filename)
            checks_data = []
            for c in val_result.checks:
                checks_data.append({
                    "Check": c.name,
                    "Category": c.category,
                    "Status": c.status,
                    "Message": c.message,
                    "Details": c.details,
                })
            df_checks = pd.DataFrame(checks_data)
            df_checks.to_excel(writer, sheet_name=sheet_name, index=False)

        # --- Changes sheet ---
        if any(r.changes for r in results):
            all_changes = []
            for r in results:
                for c in r.changes[:500]:  # Limit per file
                    all_changes.append({
                        "File": c.filename,
                        "Timestamp": c.timestamp,
                        "Type": c.field_type,
                        "Field": c.field_name,
                        "Trace": c.trace_index if c.trace_index is not None else "",
                        "Before": c.before_value,
                        "After": c.after_value,
                    })
            df_changes = pd.DataFrame(all_changes)
            df_changes.to_excel(writer, sheet_name="Changes", index=False)

    # Apply formatting
    _apply_formatting(output_path, results)

    return output_path


def write_single_validation_report(
    result: ValidationResult,
    output_path: str | Path,
) -> str:
    """Write a validation report for a single file."""
    batch = BatchResult(
        filename=result.filename,
        status=result.overall_status,
        message=f"{len(result.checks)} checks performed",
        validation_before=result,
    )
    return write_validation_report([batch], output_path)


def _apply_formatting(output_path: str, results: list[BatchResult]) -> None:
    """Apply openpyxl formatting to the report."""
    from openpyxl import load_workbook

    wb = load_workbook(output_path)

    for ws in wb.worksheets:
        # Format header row
        for cell in ws[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-width columns
        for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
            max_length = 0
            for cell in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for c in cell:
                    if c.value:
                        max_length = max(max_length, len(str(c.value)))
            adjusted = min(max_length + 4, 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

        # Color-code status cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                val = str(cell.value).upper() if cell.value else ""
                if val in _STATUS_FILLS:
                    cell.fill = _STATUS_FILLS[val]
                    cell.font = Font(bold=True)

    wb.save(output_path)


def _safe_sheet_name(filename: str) -> str:
    """Create a valid Excel sheet name from a filename."""
    name = Path(filename).stem
    # Remove invalid characters
    for ch in r"[]:*?/\\":
        name = name.replace(ch, "_")
    # Max 31 chars for Excel sheet names
    return name[:28] + "_Val" if len(name) > 28 else name + "_Val"
